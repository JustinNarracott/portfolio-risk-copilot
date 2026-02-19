"""Unit tests for file format validation (Issue #4)."""

import csv
import json
from pathlib import Path

import pytest

from src.ingestion.validators import ValidationResult, validate_file

SAMPLE_DIR = Path(__file__).parent.parent.parent / "sample-data"
FIXTURE_DIR = Path(__file__).parent.parent / "fixtures"
SAMPLE_CSV = SAMPLE_DIR / "jira-export-sample.csv"
FLAT_JSON = FIXTURE_DIR / "jira-export-flat.json"
SAMPLE_XLSX = FIXTURE_DIR / "jira-export-sample.xlsx"
NESTED_JSON = FIXTURE_DIR / "jira-export-nested.json"


# ──────────────────────────────────────────────
# File existence and extension checks
# ──────────────────────────────────────────────


class TestFileExistenceAndExtension:

    def test_missing_file(self):
        result = validate_file("/nonexistent/file.csv")
        assert result.valid is False
        assert any("not found" in e.lower() for e in result.errors)

    def test_unsupported_extension(self, tmp_path):
        f = tmp_path / "data.pdf"
        f.write_text("content")
        result = validate_file(f)
        assert result.valid is False
        assert any("unsupported" in e.lower() for e in result.errors)

    def test_empty_file(self, tmp_path):
        f = tmp_path / "empty.csv"
        f.write_text("")
        result = validate_file(f)
        assert result.valid is False
        assert any("empty" in e.lower() for e in result.errors)

    def test_returns_validation_result(self):
        result = validate_file(SAMPLE_CSV)
        assert isinstance(result, ValidationResult)

    def test_to_dict(self):
        result = validate_file(SAMPLE_CSV)
        d = result.to_dict()
        assert "valid" in d
        assert "errors" in d
        assert "warnings" in d
        assert "file_type" in d
        assert "row_count" in d


# ──────────────────────────────────────────────
# CSV validation
# ──────────────────────────────────────────────


class TestCSVValidation:

    def test_valid_sample_csv(self):
        result = validate_file(SAMPLE_CSV)
        assert result.valid is True
        assert result.errors == []
        assert result.file_type == ".csv"

    def test_row_count(self):
        result = validate_file(SAMPLE_CSV)
        assert result.row_count == 48  # 49 lines minus header

    def test_columns_found(self):
        result = validate_file(SAMPLE_CSV)
        assert "Project" in result.columns_found
        assert "Task Name" in result.columns_found

    def test_columns_mapped(self):
        result = validate_file(SAMPLE_CSV)
        assert "project" in result.columns_mapped
        assert "task_name" in result.columns_mapped
        assert "task_status" in result.columns_mapped

    def test_missing_required_columns(self, tmp_path):
        f = tmp_path / "bad.csv"
        with open(f, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["Foo", "Bar"])
            w.writerow(["val1", "val2"])
        result = validate_file(f)
        assert result.valid is False
        assert any("missing required" in e.lower() for e in result.errors)

    def test_header_only_warns(self, tmp_path):
        f = tmp_path / "headers_only.csv"
        with open(f, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["Project", "Task Name", "Task Status"])
        result = validate_file(f)
        assert result.valid is True  # Valid structure, just no data
        assert any("no data rows" in w.lower() for w in result.warnings)

    def test_minimal_columns_warns_about_optional(self, tmp_path):
        f = tmp_path / "minimal.csv"
        with open(f, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["Project", "Task Name", "Task Status"])
            w.writerow(["Alpha", "Build API", "Done"])
        result = validate_file(f)
        assert result.valid is True
        assert len(result.warnings) > 0  # Should warn about missing optional columns

    def test_optional_warning_mentions_budget(self, tmp_path):
        f = tmp_path / "no_budget.csv"
        with open(f, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["Project", "Task Name", "Task Status"])
            w.writerow(["Alpha", "Build API", "Done"])
        result = validate_file(f)
        assert any("budget" in w.lower() for w in result.warnings)

    def test_no_header_row(self, tmp_path):
        """CSV with no parseable header."""
        f = tmp_path / "noheader.csv"
        f.write_bytes(b"")  # Empty but has extension
        # Recreate with content but stat > 0
        f.write_text("\n")
        result = validate_file(f)
        assert result.valid is False

    def test_alternative_column_names_valid(self, tmp_path):
        f = tmp_path / "devops.csv"
        with open(f, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["Project Name", "Summary", "State"])
            w.writerow(["Alpha", "Build API", "Active"])
        result = validate_file(f)
        assert result.valid is True
        assert result.columns_mapped.get("project") == "Project Name"


# ──────────────────────────────────────────────
# JSON validation
# ──────────────────────────────────────────────


class TestJSONValidation:

    def test_valid_flat_json(self):
        result = validate_file(FLAT_JSON)
        assert result.valid is True
        assert result.file_type == ".json"
        assert result.row_count == 48

    def test_valid_nested_json(self):
        result = validate_file(NESTED_JSON)
        assert result.valid is True
        assert result.row_count == 48

    def test_invalid_json_syntax(self, tmp_path):
        f = tmp_path / "bad.json"
        f.write_text("{not valid json")
        result = validate_file(f)
        assert result.valid is False
        assert any("invalid json" in e.lower() for e in result.errors)

    def test_empty_list(self, tmp_path):
        f = tmp_path / "empty.json"
        f.write_text("[]")
        result = validate_file(f)
        assert result.valid is True  # Valid structure, just empty
        assert any("no data" in w.lower() for w in result.warnings)

    def test_missing_required_columns_json(self, tmp_path):
        f = tmp_path / "bad_cols.json"
        f.write_text('[{"foo": "bar"}]')
        result = validate_file(f)
        assert result.valid is False
        assert any("missing required" in e.lower() for e in result.errors)

    def test_unrecognised_structure(self, tmp_path):
        f = tmp_path / "weird.json"
        f.write_text('{"metadata": {"version": 1}}')
        result = validate_file(f)
        assert result.valid is False
        assert any("unrecognised" in e.lower() for e in result.errors)

    def test_non_object_json(self, tmp_path):
        f = tmp_path / "string.json"
        f.write_text('"just a string"')
        result = validate_file(f)
        assert result.valid is False

    def test_wrapped_json_valid(self):
        wrapped = FIXTURE_DIR / "jira-export-wrapped.json"
        result = validate_file(wrapped)
        assert result.valid is True
        assert result.row_count == 48


# ──────────────────────────────────────────────
# Excel validation
# ──────────────────────────────────────────────


class TestExcelValidation:

    def test_valid_xlsx(self):
        result = validate_file(SAMPLE_XLSX)
        assert result.valid is True
        assert result.file_type == ".xlsx"
        assert result.row_count == 48

    def test_columns_mapped_xlsx(self):
        result = validate_file(SAMPLE_XLSX)
        assert "project" in result.columns_mapped
        assert "task_name" in result.columns_mapped

    def test_empty_xlsx(self, tmp_path):
        import openpyxl
        f = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(f)
        wb.close()
        result = validate_file(f)
        # openpyxl creates one empty row by default, but no real headers
        assert result.valid is False

    def test_missing_required_columns_xlsx(self, tmp_path):
        import openpyxl
        f = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Foo", "Bar"])
        ws.append(["val1", "val2"])
        wb.save(f)
        wb.close()
        result = validate_file(f)
        assert result.valid is False
        assert any("missing required" in e.lower() for e in result.errors)

    def test_header_only_xlsx(self, tmp_path):
        import openpyxl
        f = tmp_path / "headers.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Project", "Task Name", "Task Status"])
        wb.save(f)
        wb.close()
        result = validate_file(f)
        assert result.valid is True
        assert any("no data rows" in w.lower() for w in result.warnings)

    def test_corrupt_xlsx(self, tmp_path):
        f = tmp_path / "corrupt.xlsx"
        f.write_bytes(b"this is not a real xlsx file")
        result = validate_file(f)
        assert result.valid is False
        assert any("cannot open" in e.lower() for e in result.errors)


# ──────────────────────────────────────────────
# Cross-format consistency
# ──────────────────────────────────────────────


class TestCrossFormatValidation:

    def test_all_formats_valid(self):
        for f in [SAMPLE_CSV, FLAT_JSON, SAMPLE_XLSX]:
            result = validate_file(f)
            assert result.valid is True, f"{f.name} failed validation: {result.errors}"

    def test_all_formats_same_row_count(self):
        csv_r = validate_file(SAMPLE_CSV)
        json_r = validate_file(FLAT_JSON)
        xlsx_r = validate_file(SAMPLE_XLSX)
        assert csv_r.row_count == json_r.row_count == xlsx_r.row_count
