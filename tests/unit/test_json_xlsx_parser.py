"""Unit tests for JSON and Excel parsers (Issue #3)."""

import json
from datetime import date
from pathlib import Path

import pytest

from src.ingestion.parser import Project, parse_file

# Paths to fixtures
FIXTURE_DIR = Path(__file__).parent.parent / "fixtures"
SAMPLE_DIR = Path(__file__).parent.parent.parent / "sample-data"
SAMPLE_CSV = SAMPLE_DIR / "jira-export-sample.csv"
FLAT_JSON = FIXTURE_DIR / "jira-export-flat.json"
WRAPPED_JSON = FIXTURE_DIR / "jira-export-wrapped.json"
NESTED_JSON = FIXTURE_DIR / "jira-export-nested.json"
SAMPLE_XLSX = FIXTURE_DIR / "jira-export-sample.xlsx"


@pytest.fixture()
def csv_projects() -> list[Project]:
    """Parse CSV as the reference baseline."""
    return parse_file(SAMPLE_CSV)


# ──────────────────────────────────────────────
# JSON parser — flat list structure
# ──────────────────────────────────────────────


class TestJSONFlatParser:
    """Tests for flat JSON (list of row objects)."""

    def test_returns_six_projects(self):
        projects = parse_file(FLAT_JSON)
        assert len(projects) == 6

    def test_matches_csv_project_names(self, csv_projects):
        json_projects = parse_file(FLAT_JSON)
        csv_names = [p.name for p in csv_projects]
        json_names = [p.name for p in json_projects]
        assert json_names == csv_names

    def test_matches_csv_task_counts(self, csv_projects):
        json_projects = parse_file(FLAT_JSON)
        for csv_p, json_p in zip(csv_projects, json_projects):
            assert len(json_p.tasks) == len(csv_p.tasks), f"{json_p.name} task count mismatch"

    def test_alpha_metadata(self):
        projects = parse_file(FLAT_JSON)
        alpha = next(p for p in projects if p.name == "Alpha")
        assert alpha.status == "In Progress"
        assert alpha.start_date == date(2026, 1, 6)
        assert alpha.end_date == date(2026, 6, 30)
        assert alpha.budget == 100_000.0
        assert alpha.actual_spend == 45_000.0

    def test_task_fields(self):
        projects = parse_file(FLAT_JSON)
        gamma = next(p for p in projects if p.name == "Gamma")
        migration = next(t for t in gamma.tasks if t.name == "Core platform migration")
        assert migration.status == "In Progress"
        assert migration.priority == "Critical"
        assert migration.previous_sprints == ["Sprint 4", "Sprint 5", "Sprint 6", "Sprint 7"]


# ──────────────────────────────────────────────
# JSON parser — wrapped dict structure
# ──────────────────────────────────────────────


class TestJSONWrappedParser:
    """Tests for wrapped JSON (dict with 'issues' key)."""

    def test_returns_six_projects(self):
        projects = parse_file(WRAPPED_JSON)
        assert len(projects) == 6

    def test_matches_csv_project_names(self, csv_projects):
        json_projects = parse_file(WRAPPED_JSON)
        csv_names = [p.name for p in csv_projects]
        json_names = [p.name for p in json_projects]
        assert json_names == csv_names

    def test_matches_csv_task_counts(self, csv_projects):
        json_projects = parse_file(WRAPPED_JSON)
        for csv_p, json_p in zip(csv_projects, json_projects):
            assert len(json_p.tasks) == len(csv_p.tasks), f"{json_p.name} task count mismatch"


# ──────────────────────────────────────────────
# JSON parser — nested project structure
# ──────────────────────────────────────────────


class TestJSONNestedParser:
    """Tests for nested JSON (projects with embedded tasks)."""

    def test_returns_six_projects(self):
        projects = parse_file(NESTED_JSON)
        assert len(projects) == 6

    def test_matches_csv_project_names(self, csv_projects):
        json_projects = parse_file(NESTED_JSON)
        csv_names = [p.name for p in csv_projects]
        json_names = [p.name for p in json_projects]
        assert json_names == csv_names

    def test_matches_csv_task_counts(self, csv_projects):
        json_projects = parse_file(NESTED_JSON)
        for csv_p, json_p in zip(csv_projects, json_projects):
            assert len(json_p.tasks) == len(csv_p.tasks), f"{json_p.name} task count mismatch"

    def test_gamma_metadata(self):
        projects = parse_file(NESTED_JSON)
        gamma = next(p for p in projects if p.name == "Gamma")
        assert gamma.status == "At Risk"
        assert gamma.budget == 200_000.0
        assert gamma.actual_spend == 185_000.0


# ──────────────────────────────────────────────
# JSON parser — edge cases
# ──────────────────────────────────────────────


class TestJSONEdgeCases:
    """Tests for JSON edge cases and error handling."""

    def test_empty_list(self, tmp_path):
        f = tmp_path / "empty.json"
        f.write_text("[]")
        projects = parse_file(f)
        assert projects == []

    def test_empty_dict_with_tasks(self, tmp_path):
        f = tmp_path / "empty_tasks.json"
        f.write_text('{"tasks": []}')
        projects = parse_file(f)
        assert projects == []

    def test_missing_required_columns(self, tmp_path):
        f = tmp_path / "bad.json"
        f.write_text('[{"foo": "bar", "baz": 123}]')
        with pytest.raises(ValueError, match="Missing required columns"):
            parse_file(f)

    def test_unrecognised_structure(self, tmp_path):
        f = tmp_path / "weird.json"
        f.write_text('{"metadata": {"version": 1}}')
        with pytest.raises(ValueError, match="Unrecognised JSON structure"):
            parse_file(f)

    def test_null_values_handled(self, tmp_path):
        f = tmp_path / "nulls.json"
        data = [
            {"Project": "Alpha", "Task Name": "Build API", "Task Status": "Done",
             "Budget": None, "Comments": None, "Previous Sprints": None}
        ]
        f.write_text(json.dumps(data))
        projects = parse_file(f)
        assert len(projects) == 1
        assert projects[0].budget == 0.0
        assert projects[0].tasks[0].comments == ""
        assert projects[0].tasks[0].previous_sprints == []

    def test_numeric_values_stringified(self, tmp_path):
        f = tmp_path / "nums.json"
        data = [
            {"Project": "Alpha", "Task Name": "Build API", "Task Status": "Done",
             "Budget": 100000, "Actual Spend": 45000.5}
        ]
        f.write_text(json.dumps(data))
        projects = parse_file(f)
        assert projects[0].budget == 100_000.0
        assert projects[0].actual_spend == 45_000.5

    def test_list_values_joined(self, tmp_path):
        """Lists (e.g., previous_sprints as array) are joined with semicolons."""
        f = tmp_path / "lists.json"
        data = [
            {"Project": "Alpha", "Task Name": "Build API", "Task Status": "Done",
             "Previous Sprints": ["Sprint 3", "Sprint 4", "Sprint 5"]}
        ]
        f.write_text(json.dumps(data))
        projects = parse_file(f)
        assert projects[0].tasks[0].previous_sprints == ["Sprint 3", "Sprint 4", "Sprint 5"]


# ──────────────────────────────────────────────
# Excel parser tests
# ──────────────────────────────────────────────


class TestExcelParser:
    """Tests for XLSX parser."""

    def test_returns_six_projects(self):
        projects = parse_file(SAMPLE_XLSX)
        assert len(projects) == 6

    def test_matches_csv_project_names(self, csv_projects):
        xlsx_projects = parse_file(SAMPLE_XLSX)
        csv_names = [p.name for p in csv_projects]
        xlsx_names = [p.name for p in xlsx_projects]
        assert xlsx_names == csv_names

    def test_matches_csv_task_counts(self, csv_projects):
        xlsx_projects = parse_file(SAMPLE_XLSX)
        for csv_p, xlsx_p in zip(csv_projects, xlsx_projects):
            assert len(xlsx_p.tasks) == len(csv_p.tasks), f"{xlsx_p.name} task count mismatch"

    def test_alpha_metadata(self):
        projects = parse_file(SAMPLE_XLSX)
        alpha = next(p for p in projects if p.name == "Alpha")
        assert alpha.status == "In Progress"
        assert alpha.budget == 100_000.0
        assert alpha.actual_spend == 45_000.0

    def test_task_fields(self):
        projects = parse_file(SAMPLE_XLSX)
        alpha = next(p for p in projects if p.name == "Alpha")
        blocked = next(t for t in alpha.tasks if t.name == "Implement payment gateway")
        assert blocked.status == "Blocked"
        assert blocked.priority == "Critical"
        assert "Blocked by third-party" in blocked.comments

    def test_previous_sprints(self):
        projects = parse_file(SAMPLE_XLSX)
        gamma = next(p for p in projects if p.name == "Gamma")
        migration = next(t for t in gamma.tasks if t.name == "Core platform migration")
        assert migration.previous_sprints == ["Sprint 4", "Sprint 5", "Sprint 6", "Sprint 7"]


class TestExcelEdgeCases:
    """Tests for Excel edge cases."""

    def test_empty_xlsx(self, tmp_path):
        import openpyxl
        f = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(f)
        wb.close()
        with pytest.raises(ValueError, match="empty"):
            parse_file(f)

    def test_missing_required_columns_xlsx(self, tmp_path):
        import openpyxl
        f = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Foo", "Bar"])
        ws.append(["val1", "val2"])
        wb.save(f)
        wb.close()
        with pytest.raises(ValueError, match="Missing required columns"):
            parse_file(f)

    def test_numeric_cells_handled(self, tmp_path):
        """Excel numeric cells (int/float) are converted to strings correctly."""
        import openpyxl
        f = tmp_path / "nums.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Project", "Task Name", "Task Status", "Budget", "Actual Spend"])
        ws.append(["Alpha", "Build API", "Done", 100000, 45000.5])
        wb.save(f)
        wb.close()

        projects = parse_file(f)
        assert projects[0].budget == 100_000.0
        assert projects[0].actual_spend == 45_000.5

    def test_date_cells_handled(self, tmp_path):
        """Excel date cells are converted to YYYY-MM-DD strings."""
        import openpyxl
        from datetime import datetime

        f = tmp_path / "dates.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Project", "Task Name", "Task Status", "Start Date", "End Date"])
        ws.append(["Alpha", "Build API", "Done", datetime(2026, 1, 6), datetime(2026, 6, 30)])
        wb.save(f)
        wb.close()

        projects = parse_file(f)
        assert projects[0].start_date == date(2026, 1, 6)
        assert projects[0].end_date == date(2026, 6, 30)


# ──────────────────────────────────────────────
# Cross-format consistency test
# ──────────────────────────────────────────────


class TestCrossFormatConsistency:
    """Verify all three formats produce identical output."""

    def test_all_formats_same_projects(self):
        csv_p = parse_file(SAMPLE_CSV)
        json_p = parse_file(FLAT_JSON)
        xlsx_p = parse_file(SAMPLE_XLSX)

        csv_names = [p.name for p in csv_p]
        json_names = [p.name for p in json_p]
        xlsx_names = [p.name for p in xlsx_p]

        assert csv_names == json_names == xlsx_names

    def test_all_formats_same_task_counts(self):
        csv_p = parse_file(SAMPLE_CSV)
        json_p = parse_file(FLAT_JSON)
        xlsx_p = parse_file(SAMPLE_XLSX)

        for c, j, x in zip(csv_p, json_p, xlsx_p):
            assert len(c.tasks) == len(j.tasks) == len(x.tasks), (
                f"{c.name}: CSV={len(c.tasks)}, JSON={len(j.tasks)}, XLSX={len(x.tasks)}"
            )

    def test_all_formats_same_budgets(self):
        csv_p = parse_file(SAMPLE_CSV)
        json_p = parse_file(FLAT_JSON)
        xlsx_p = parse_file(SAMPLE_XLSX)

        for c, j, x in zip(csv_p, json_p, xlsx_p):
            assert c.budget == j.budget == x.budget, f"{c.name} budget mismatch"
            assert c.actual_spend == j.actual_spend == x.actual_spend, f"{c.name} spend mismatch"
